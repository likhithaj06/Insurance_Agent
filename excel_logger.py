"""
excel_logger.py
---------------
Logs every processed claim into an Excel file (claims_log.xlsx), one row per
policy number, so the agent's decisions can be reviewed or audited later.

Behavior:
  - If a policy number already has a row (i.e. the same claim was processed
    again), that row is UPDATED in place rather than duplicated.
  - If the policy number is new (or missing), a new row is appended.
  - The workbook is created automatically on first run if it doesn't exist.

This is intentionally isolated in its own module: if Excel logging ever
fails (file locked, disk full, permissions issue), it must NEVER break the
main agent pipeline. Callers should wrap calls to this module in try/except,
which pipeline.py already does.
"""

from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LOG_FILE_PATH = "claims_log.xlsx"
SHEET_NAME = "Claims Log"

# Column order for the log -- mirrors schema.MANDATORY_FIELDS plus the
# routing outcome fields and a processed-at timestamp for traceability.
COLUMNS = [
    "policy_number", "policyholder_name", "effective_dates",
    "date", "time", "location", "description",
    "claimant", "third_parties", "contact_details",
    "asset_type", "asset_id", "estimated_damage",
    "claim_type", "attachments", "initial_estimate",
    "missing_fields", "recommended_route", "reasoning",
    "extraction_method", "processed_at",
]

HEADER_LABELS = {
    "policy_number": "Policy Number",
    "policyholder_name": "Policyholder Name",
    "effective_dates": "Effective Dates",
    "date": "Incident Date",
    "time": "Incident Time",
    "location": "Location",
    "description": "Description",
    "claimant": "Claimant",
    "third_parties": "Third Parties",
    "contact_details": "Contact Details",
    "asset_type": "Asset Type",
    "asset_id": "Asset ID",
    "estimated_damage": "Estimated Damage",
    "claim_type": "Claim Type",
    "attachments": "Attachments",
    "initial_estimate": "Initial Estimate",
    "missing_fields": "Missing Fields",
    "recommended_route": "Recommended Route",
    "reasoning": "Reasoning",
    "extraction_method": "Extraction Method",
    "processed_at": "Processed At",
}


def _create_new_workbook():
    """Builds a fresh workbook with a styled header row."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = SHEET_NAME

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")

    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=HEADER_LABELS[col_key])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"  # keep header visible when scrolling
    return wb


def _find_row_for_policy(sheet, policy_number: str) -> int | None:
    """Returns the row number (1-indexed) of an existing entry for this policy, or None."""
    if not policy_number:
        return None
    policy_col_idx = COLUMNS.index("policy_number") + 1
    for row_idx in range(2, sheet.max_row + 1):
        existing_value = sheet.cell(row=row_idx, column=policy_col_idx).value
        if existing_value == policy_number:
            return row_idx
    return None


def _write_row(sheet, row_idx: int, row_values: dict):
    """Writes/overwrites a single row with values from row_values (keyed by column name)."""
    body_font = Font(name="Arial", size=10)
    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=row_values.get(col_key, ""))
        cell.font = body_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize_columns(sheet):
    """Sets a reasonable fixed width per column (autosize-by-content is unreliable with wrap_text)."""
    narrow_cols = {"date", "time", "policy_number", "extraction_method", "processed_at"}
    wide_cols = {"description", "reasoning", "contact_details", "missing_fields"}
    for col_idx, col_key in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if col_key in wide_cols:
            sheet.column_dimensions[letter].width = 40
        elif col_key in narrow_cols:
            sheet.column_dimensions[letter].width = 16
        else:
            sheet.column_dimensions[letter].width = 22


def log_claim(agent_output: dict, file_path: str = LOG_FILE_PATH):
    """
    Appends or updates a row in the Excel log for the given agent output
    (the same dict returned by pipeline.run_fnol_agent).

    Upsert key: policy_number. If it's missing/None, the row is always
    appended (there's no reliable key to match against).
    """
    fields = agent_output.get("extractedFields", {})
    policy_number = fields.get("policy_number")

    row_values = dict(fields)  # policy_number..initial_estimate
    row_values["missing_fields"] = ", ".join(agent_output.get("missingFields", [])) or "None"
    row_values["recommended_route"] = agent_output.get("recommendedRoute", "")
    row_values["reasoning"] = agent_output.get("reasoning", "")
    row_values["extraction_method"] = agent_output.get("_extractionMethod", "unknown")
    row_values["processed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        wb = load_workbook(file_path)
        sheet = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    except FileNotFoundError:
        wb = _create_new_workbook()
        sheet = wb[SHEET_NAME]

    existing_row = _find_row_for_policy(sheet, policy_number)
    target_row = existing_row if existing_row else sheet.max_row + 1
    _write_row(sheet, target_row, row_values)
    _autosize_columns(sheet)

    wb.save(file_path)
    return "updated" if existing_row else "appended"
